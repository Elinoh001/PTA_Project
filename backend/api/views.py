import logging
import io
from datetime import datetime
from django.db.models import Count, Sum, Avg
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.styles import numbers
from django.contrib.auth.models import User
from .models import UserProfile, Service, Activite, PCOPEntry, Suivi, ObjectifGeneral, ObjectifSpecifique, ResultatAttendu, Direction, Division, Structure
from .serializers import UserProfileSerializer, ServiceSerializer, ActiviteSerializer, PCOPEntrySerializer, SuiviSerializer, ObjectifGeneralSerializer, ObjectifSpecifiqueSerializer, ResultatAttenduSerializer, DirectionSerializer, DivisionSerializer, StructureSerializer
from .permissions import RolePermission, AdminOnlyPermission, SuperviseurAndAdminPermission, ReadOnlyPermission

# Configuration du logger
logger = logging.getLogger(__name__)

# ViewSets avec permissions spécifiques
class UserProfileViewSet(viewsets.ModelViewSet):
    queryset = UserProfile.objects.all()
    serializer_class = UserProfileSerializer
    permission_classes = [IsAuthenticated, AdminOnlyPermission]

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            return UserProfile.objects.all()
        try:
            user_profile = user.userprofile
            if user_profile.role == 'admin':
                return UserProfile.objects.all()
            elif user_profile.role == 'superviseur':
                return UserProfile.objects.exclude(role='admin')
            elif user_profile.role == 'user':
                return UserProfile.objects.filter(auth_user=user)
        except AttributeError:
            pass
        return UserProfile.objects.none()

# ✅ NOUVEAUX VIEWSETS POUR LA STRUCTURE ORGANISATIONNELLE HIÉRARCHIQUE
class StructureViewSet(viewsets.ModelViewSet):
    queryset = Structure.objects.prefetch_related('directions__services__divisions').all()
    serializer_class = StructureSerializer
    permission_classes = [IsAuthenticated, SuperviseurAndAdminPermission]

class DirectionViewSet(viewsets.ModelViewSet):
    queryset = Direction.objects.select_related('structure').prefetch_related('services__divisions').all()
    serializer_class = DirectionSerializer
    permission_classes = [IsAuthenticated, SuperviseurAndAdminPermission]

class ServiceViewSet(viewsets.ModelViewSet):
    queryset = Service.objects.select_related('direction__structure').prefetch_related('divisions').all()
    serializer_class = ServiceSerializer
    permission_classes = [IsAuthenticated, SuperviseurAndAdminPermission]

class DivisionViewSet(viewsets.ModelViewSet):
    queryset = Division.objects.select_related('service__direction__structure').all()
    serializer_class = DivisionSerializer
    permission_classes = [IsAuthenticated, SuperviseurAndAdminPermission]

# ✅ VIEWSETS POUR LA STRUCTURE HIÉRARCHIQUE DES OBJECTIFS
class ObjectifGeneralViewSet(viewsets.ModelViewSet):
    queryset = ObjectifGeneral.objects.prefetch_related('objectifs_specifiques__resultats_attendus').all()
    serializer_class = ObjectifGeneralSerializer
    permission_classes = [IsAuthenticated, RolePermission]

class ObjectifSpecifiqueViewSet(viewsets.ModelViewSet):
    queryset = ObjectifSpecifique.objects.select_related('objectif_general').prefetch_related('resultats_attendus').all()
    serializer_class = ObjectifSpecifiqueSerializer
    permission_classes = [IsAuthenticated, RolePermission]

class ResultatAttenduViewSet(viewsets.ModelViewSet):
    queryset = ResultatAttendu.objects.select_related('objectif_specifique__objectif_general').all()
    serializer_class = ResultatAttenduSerializer
    permission_classes = [IsAuthenticated, RolePermission]

class ActiviteViewSet(viewsets.ModelViewSet):
    queryset = Activite.objects.select_related(
        'structure',
        'direction', 
        'service', 
        'division',
        'objectif_general', 
        'objectif_specifique', 
        'resultat_attendu',
        'pcop'
    ).all()
    serializer_class = ActiviteSerializer
    permission_classes = [IsAuthenticated, RolePermission]

    def get_queryset(self):
        return Activite.objects.select_related(
            'structure',
            'direction',
            'service', 
            'division',
            'objectif_general', 
            'objectif_specifique', 
            'resultat_attendu',
            'pcop'
        ).prefetch_related('suivis')

    def perform_create(self, serializer):
        data = serializer.validated_data
        cout_unitaire = data.get('cout_unitaire')
        quantite = data.get('quantite')
        
        if cout_unitaire is not None and quantite is not None:
            data['montant'] = cout_unitaire * quantite
        
        serializer.save()

class PCOPEntryViewSet(viewsets.ModelViewSet):
    queryset = PCOPEntry.objects.all()
    serializer_class = PCOPEntrySerializer
    permission_classes = [IsAuthenticated, SuperviseurAndAdminPermission]

class SuiviViewSet(viewsets.ModelViewSet):
    queryset = Suivi.objects.select_related('activite').all()
    serializer_class = SuiviSerializer
    permission_classes = [IsAuthenticated, RolePermission]

    def get_queryset(self):
        user = self.request.user
        queryset = Suivi.objects.select_related('activite')
        
        activite_id = self.request.query_params.get('activite_id')
        if activite_id:
            queryset = queryset.filter(activite_id=activite_id)
        
        date_suivi = self.request.query_params.get('date_suivi')
        if date_suivi:
            queryset = queryset.filter(date_suivi=date_suivi)
        
        if user.is_superuser:
            return queryset
        try:
            user_profile = user.userprofile
            if user_profile.role in ['admin', 'superviseur']:
                return queryset
        except AttributeError:
            pass
        return queryset

    def perform_create(self, serializer):
        data = serializer.validated_data
        avancement = data.get('avancement', 0)
        
        if avancement < 0 or avancement > 100:
            from rest_framework import serializers
            raise serializers.ValidationError({
                'avancement': 'L\'avancement doit être entre 0 et 100%'
            })
        
        serializer.save()

# VUES EXISTANTES (inchangées)
@api_view(['POST'])
@permission_classes([IsAuthenticated, AdminOnlyPermission])
def create_user_with_profile(request):
    try:
        data = request.data
        required_fields = ['nom', 'email', 'role']
        for field in required_fields:
            if not data.get(field):
                return Response(
                    {'error': f'Le champ {field} est obligatoire'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        if User.objects.filter(email=data['email']).exists():
            return Response(
                {'error': 'Un utilisateur avec cet email existe déjà'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user = User.objects.create_user(
            username=data['email'],
            email=data['email'],
            password=data.get('password', 'password123'),
            first_name=data.get('nom', '')
        )
        
        profile = UserProfile.objects.create(
            nom=data['nom'],
            email=data['email'],
            role=data.get('role', 'user'),
            auth_user=user
        )
        
        serializer = UserProfileSerializer(profile)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return Response(
            {'error': f'Erreur lors de la création: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, AdminOnlyPermission])
def update_user_role(request, user_id):
    try:
        profile = UserProfile.objects.get(id=user_id)
        new_role = request.data.get('role')
        
        if new_role not in ['admin', 'superviseur', 'user']:
            return Response(
                {'error': 'Rôle invalide'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        profile.role = new_role
        profile.save()
        
        serializer = UserProfileSerializer(profile)
        return Response(serializer.data)
        
    except UserProfile.DoesNotExist:
        return Response(
            {'error': 'Utilisateur non trouvé'},
            status=status.HTTP_404_NOT_FOUND
        )

@api_view(['POST'])
@permission_classes([IsAuthenticated, AdminOnlyPermission])
def reset_user_password(request, user_id):
    try:
        profile = UserProfile.objects.get(id=user_id)
        user = profile.auth_user
        
        if not user:
            return Response(
                {'error': 'Utilisateur Django non trouvé'},
                status=status.HTTP_404_NOT_FOUND
            )
            
        new_password = request.data.get('new_password', 'password123')
        user.set_password(new_password)
        user.save()
        
        return Response({'message': 'Mot de passe réinitialisé avec succès'})
        
    except UserProfile.DoesNotExist:
        return Response(
            {'error': 'Utilisateur non trouvé'},
            status=status.HTTP_404_NOT_FOUND
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_profile(request):
    try:
        user_profile = UserProfile.objects.get(auth_user=request.user)
        serializer = UserProfileSerializer(user_profile)
        return Response(serializer.data)
        
    except UserProfile.DoesNotExist:
        user_profile = UserProfile.objects.create(
            nom=request.user.first_name or request.user.username,
            email=request.user.email,
            role='user',
            auth_user=request.user
        )
        serializer = UserProfileSerializer(user_profile)
        return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated, SuperviseurAndAdminPermission])
def get_dashboard_stats(request):
    stats = {
        'total_users': UserProfile.objects.count(),
        'total_directions': Direction.objects.count(),
        'total_services': Service.objects.count(),
        'total_divisions': Division.objects.count(),
        'total_activites': Activite.objects.count(),
        'total_suivis': Suivi.objects.count(),
        'total_structures': Structure.objects.count(),  # Correction: Count() -> count()
        'budget_total': PCOPEntry.objects.aggregate(total=Sum('budget'))['total'] or 0,
        'montant_total_activites': Activite.objects.aggregate(total=Sum('montant'))['total'] or 0,
        'moyenne_avancement': Suivi.objects.aggregate(moyenne=Avg('avancement'))['moyenne'] or 0,
        'users_by_role': {
            'admin': UserProfile.objects.filter(role='admin').count(),
            'superviseur': UserProfile.objects.filter(role='superviseur').count(),
            'user': UserProfile.objects.filter(role='user').count(),
        },
        'activites_by_etat': {
            'en_cours': Activite.objects.filter(etat='En cours').count(),
            'termine': Activite.objects.filter(etat='Terminé').count(),
            'en_attente': Activite.objects.filter(etat='En attente').count(),
        },
        # ✅ NOUVELLES STATISTIQUES
        'objectifs_generaux_count': ObjectifGeneral.objects.count(),
        'objectifs_specifiques_count': ObjectifSpecifique.objects.count(),
        'resultats_attendus_count': ResultatAttendu.objects.count(),
        'activites_by_structure': {
            structure.nom: Activite.objects.filter(structure=structure).count()
            for structure in Structure.objects.all()
        }
    }
    return Response(stats)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_activites_by_service(request):
    services = Service.objects.annotate(
        nb_activites=Count('activites')
    ).values('id', 'nom_service', 'nb_activites')
    
    return Response(list(services))

# ✅ NOUVELLE VUE POUR RÉCUPÉRER TOUTES LES DONNÉES INITIALES
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_initial_data(request):
    """Endpoint pour récupérer toutes les données initiales nécessaires au frontend"""
    try:
        data = {
            'structures': StructureSerializer(
                Structure.objects.prefetch_related('directions__services__divisions').all(), 
                many=True
            ).data,
            'objectifs_generaux': ObjectifGeneralSerializer(
                ObjectifGeneral.objects.prefetch_related('objectifs_specifiques__resultats_attendus').all(), 
                many=True
            ).data,
            'pcop_entries': PCOPEntrySerializer(
                PCOPEntry.objects.all(), 
                many=True
            ).data,
        }
        return Response(data)
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des données initiales: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ✅ MODIFICATION DE L'EXPORT EXCEL POUR INCLURE LA NOUVELLE STRUCTURE
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_pta_excel(request):
    try:
        logger.info(f"Début de l'export Excel par l'utilisateur: {request.user.username}")
        
        wb = openpyxl.Workbook()
        
        # Styles réutilisables
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        border_style = Side(border_style="thin", color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # ---------- FEUILLE PRINCIPALE PTA ----------
        ws_pta = wb.active
        ws_pta.title = "PTA_PRINCIPAL"
        
        ws_pta.merge_cells('A1:S1')
        title_cell = ws_pta.cell(row=1, column=1)
        title_cell.value = "PLAN DE TRAVAIL ANNUEL (PTA) - EXPORT COMPLET"
        title_cell.font = Font(bold=True, size=16, color="2E86AB")
        title_cell.alignment = center_align
        
        ws_pta.merge_cells('A2:S2')
        meta_cell = ws_pta.cell(row=2, column=1)
        meta_cell.value = f"Export généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} par {request.user.username}"
        meta_cell.font = Font(italic=True, size=10, color="666666")
        meta_cell.alignment = center_align
        
        headers = [
            "OBJECTIFS GENERAUX", "OBJECTIFS SPECIFIQUES", "RESULTATS ATTENDUS",
            "STRUCTURE", "DIRECTION", "SERVICE", "DIVISION", "ACTIVITES",
            "SOUS-ACTIVITES", "PRODUITS", "CIBLES", "SOURCES DE FINANCEMENT",
            "CODE PCOP", "LIBELLE PCOP", "COUT UNITAIRE (Ar)", "QUANTITE", "MONTANT TOTAL (Ar)", "OBSERVATIONS", "ETAT"
        ]
        
        ws_pta.append([])
        ws_pta.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_pta.cell(row=4, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # ✅ MODIFICATION: Inclure toutes les relations hiérarchiques
        activites = Activite.objects.select_related(
            'structure',
            'direction', 
            'service', 
            'division',
            'objectif_general', 
            'objectif_specifique', 
            'resultat_attendu',
            'pcop'
        ).all()
        
        total_montant = 0
        row_count = 0
        
        for activite in activites:
            # ✅ MODIFICATION: Récupérer les informations via la nouvelle structure
            row_data = [
                activite.objectif_general_titre if activite.objectif_general else "Non spécifié",
                activite.objectif_specifique_titre if activite.objectif_specifique else "Non spécifié",
                activite.resultat_attendu_description if activite.resultat_attendu else "Non spécifié",
                activite.structure_nom if activite.structure else "Non spécifié",
                activite.direction_nom if activite.direction else "Non spécifié",
                activite.service_nom if activite.service else "Non assigné",
                activite.division_nom if activite.division else "Non spécifié",
                activite.activite or "Non spécifié",
                activite.sous_activite or "Non spécifié",
                activite.produits or "Non spécifié",
                activite.cibles or "Non spécifié",
                activite.sources_financement or "Non spécifié",
                activite.pcop_code if activite.pcop else "Non spécifié",
                activite.pcop_libelle if activite.pcop else "Non spécifié",
                float(activite.cout_unitaire) if activite.cout_unitaire else 0.0,
                float(activite.quantite) if activite.quantite else 0.0,
                float(activite.montant) if activite.montant else 0.0,
                activite.observation or "Aucune",
                activite.etat or "En cours"
            ]
            
            ws_pta.append(row_data)
            row_count += 1
            
            if activite.montant:
                total_montant += activite.montant
            
            for col_num in range(1, len(headers) + 1):
                cell = ws_pta.cell(row=4 + row_count, column=col_num)
                cell.border = border
                cell.alignment = left_align
                
                if col_num in [15, 16, 17]:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
        
        if row_count > 0:
            total_row = 5 + row_count
            ws_pta.merge_cells(f'A{total_row}:P{total_row}')
            total_label = ws_pta.cell(row=total_row, column=1)
            total_label.value = f"TOTAL GÉNÉRAL ({row_count} activités)"
            total_label.font = Font(bold=True, size=12, color="2E86AB")
            total_label.alignment = Alignment(horizontal="right", vertical="center")
            
            total_cell = ws_pta.cell(row=total_row, column=17)
            total_cell.value = float(total_montant)
            total_cell.font = Font(bold=True, size=12, color="2E86AB")
            total_cell.number_format = '#,##0.00'
            total_cell.border = border
        
        for column in ws_pta.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws_pta.column_dimensions[column_letter].width = adjusted_width
        
        # ---------- FEUILLE STRUCTURE LOGIQUE ----------
        ws_structure = wb.create_sheet("STRUCTURE_LOGIQUE")
        
        ws_structure.merge_cells('A1:D1')
        title_cell = ws_structure.cell(row=1, column=1)
        title_cell.value = "STRUCTURE LOGIQUE - OBJECTIFS ET RÉSULTATS"
        title_cell.font = Font(bold=True, size=14, color="2E86AB")
        title_cell.alignment = center_align
        
        structure_headers = ["Objectif Général", "Objectif Spécifique", "Résultat Attendu", "Nombre d'Activités"]
        ws_structure.append(structure_headers)
        
        for col_num, header in enumerate(structure_headers, 1):
            cell = ws_structure.cell(row=2, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        row_num = 3
        objectifs_generaux = ObjectifGeneral.objects.prefetch_related(
            'objectifs_specifiques__resultats_attendus'
        ).all()
        
        for og in objectifs_generaux:
            for os in og.objectifs_specifiques.all():
                for ra in os.resultats_attendus.all():
                    activites_count = Activite.objects.filter(resultat_attendu=ra).count()
                    ws_structure.append([
                        f"{og.numero} - {og.titre}",
                        f"{os.numero} - {os.titre}",
                        f"{ra.numero} - {ra.description}",
                        activites_count
                    ])
                    
                    for col_num in range(1, 5):
                        cell = ws_structure.cell(row=row_num, column=col_num)
                        cell.border = border
                        cell.alignment = left_align
                    
                    row_num += 1
        
        # Ajustement des colonnes
        for column in ws_structure.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws_structure.column_dimensions[column_letter].width = adjusted_width
        
        # ---------- FEUILLE STRUCTURE ORGANISATIONNELLE ----------
        ws_org = wb.create_sheet("STRUCTURE_ORGANISATIONNELLE")
        
        ws_org.merge_cells('A1:E1')
        title_cell = ws_org.cell(row=1, column=1)
        title_cell.value = "STRUCTURE ORGANISATIONNELLE - STRUCTURES, DIRECTIONS, SERVICES ET DIVISIONS"
        title_cell.font = Font(bold=True, size=14, color="2E86AB")
        title_cell.alignment = center_align
        
        org_headers = ["Structure", "Direction", "Service", "Division", "Nombre d'Activités"]
        ws_org.append(org_headers)
        
        for col_num, header in enumerate(org_headers, 1):
            cell = ws_org.cell(row=2, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        row_num = 3
        structures = Structure.objects.prefetch_related('directions__services__divisions').all()
        
        for structure in structures:
            for direction in structure.directions.all():
                for service in direction.services.all():
                    for division in service.divisions.all():
                        activites_count = Activite.objects.filter(division=division).count()
                        ws_org.append([
                            f"{structure.numero} - {structure.nom}",
                            f"{direction.numero} - {direction.nom}",
                            f"{service.numero} - {service.nom_service}",
                            f"{division.numero} - {division.nom}",
                            activites_count
                        ])
                    
                        for col_num in range(1, 6):
                            cell = ws_org.cell(row=row_num, column=col_num)
                            cell.border = border
                            cell.alignment = left_align
                    
                        row_num += 1
        
        # Ajustement des colonnes
        for column in ws_org.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws_org.column_dimensions[column_letter].width = adjusted_width
        
        # ---------- PRÉPARATION DE LA RÉPONSE ----------
        file_buffer = io.BytesIO()
        wb.save(file_buffer)
        file_buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PTA_Export_Complet_{timestamp}.xlsx"
        
        response = HttpResponse(
            file_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(file_buffer.getvalue())
        
        logger.info(f"Export Excel réussi: {filename} - {row_count} activités exportées")
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'export Excel: {str(e)}", exc_info=True)
        error_message = f"Erreur lors de l'export Excel: {str(e)}"
        return Response({'error': error_message}, status=500)